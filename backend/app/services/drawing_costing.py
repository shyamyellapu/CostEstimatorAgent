"""
Drawing Costing Service
-----------------------
Handles the full pipeline for Job Costing Sheet automation from a drawing PDF:
  1. Calls Claude vision API with the structural steel takeoff prompt
  2. Computes steel weight from extracted members + plates
  3. Computes all derived quantities and cost line items
  4. Resolves the D58 ↔ J33 circular reference
  5. Stamps computed values into the Sample Job Costing Sheet.xlsx template
     (preserving all formatting, merges, and pre-built formulas)
"""
import base64
import json
import logging
import math
import re
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

import anthropic
import openpyxl

from app.config import settings

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Section-weight lookup table (kg/m)
# ---------------------------------------------------------------------------
SECTION_WEIGHTS: Dict[str, float] = {
    # Universal Columns
    "UC 152x152x23": 23.0,  "UC 152x152x30": 30.0,  "UC 152x152x37": 37.0,
    "UC 203x203x46": 46.1,  "UC 203x203x52": 52.0,  "UC 203x203x60": 60.0,
    # Universal Beams
    "UB 127x76x13":  13.0,  "UB 152x89x16":  16.0,  "UB 178x102x19": 19.0,
    "UB 203x102x23": 23.1,  "UB 203x133x25": 25.1,  "UB 203x133x30": 30.0,
    # Split tees
    "UCT 152x152x30": 30.0, "UBT 133x101x15": 15.0, "UBT 133x102x15": 15.0,
    # Parallel Flange Channels
    "PFC 100x50x10": 10.2,  "PFC 125x65x15": 14.8,  "PFC 150x75x18": 17.9,
    "PFC 150x90x24": 23.9,  "PFC 180x90x26": 26.1,  "PFC 200x75x23": 23.4,
    "PFC 230x75x26": 25.7,  "PFC 230x90x32": 32.2,
    # Equal Angles
    "L 50x50x6":  4.47,  "L 75x75x8":  8.99,
    "L 100x100x10": 15.0, "L 100x100x12": 17.8, "L 150x150x12": 27.3,
}

PLATE_DENSITY = 7.85   # kg per m² per mm thickness

# ---------------------------------------------------------------------------
# Derived-ratio constants (reverse-engineered from drawing 1349001-B)
# ---------------------------------------------------------------------------
RATIOS: Dict[str, float] = {
    "paintLitresPerSqm":   0.6,
    "surfaceAreaPerKg":    0.02564,
    "boltsPer1000Kg":      3.85,
    "mpiVisitsPer1000Kg":  1.28,
    "weldingMHPerKg":      0.02050,
    "fabricationMHPerKg":  0.04101,
    # Unit rates (AED)
    "rateSteelPerKg":       4.0,
    "rateBoltPerNo":        12.5,
    "ratePaintPerLitre":    21.0,
    "rateWeldingPerMH":     10.5,
    "rateFabricationPerMH": 9.5,
    "rateBlastingPerSqm":   9.0,
    "ratePaintingPerSqm":   11.0,
    "rateMPIPerVisit":      600.0,
    "rateQAQC":             3000.0,
    "ratePacking":          3000.0,
    "defaultMarkupPct":     0.34,
}

# ---------------------------------------------------------------------------
# Vision extraction prompt (verbatim — DO NOT MODIFY)
# ---------------------------------------------------------------------------
EXTRACTION_PROMPT = """You are a structural steel takeoff engineer analyzing engineering drawings for an Oil & Gas pipe support package. Extract ALL structural members from every page.

For EACH unique member type you see (e.g., "UC 152x152x30", "PFC 150x90x24", "L 100x100x10", "UB 203x133x30", "UBT 133x101x15"), identify:
1. The exact section designation (preserve formatting like "UC 152x152x30")
2. Total estimated linear length in METERS across ALL pipe support details on ALL pages (sum all instances). Use the dimensions shown in the drawing (mm → convert to m).
3. Approximate count of pieces

Also identify:
- Any plates: thickness (mm), approximate total area (m²), and description
- Bracing angles separately from columns/beams
- Drawing/project metadata: project name, drawing number, revision, client, contractor

Return ONLY valid JSON with this exact shape (no markdown fences, no commentary):
{
  "project": {
    "drawing_no": "string",
    "revision": "string",
    "title": "string",
    "client": "string",
    "contractor": "string",
    "package_description": "string"
  },
  "members": [
    {"section": "UC 152x152x30", "total_length_m": 24.5, "pieces": 12, "role": "column"},
    {"section": "PFC 150x90x24", "total_length_m": 18.0, "pieces": 8, "role": "beam"}
  ],
  "plates": [
    {"description": "Cap Plate", "thickness_mm": 10, "total_area_m2": 0.45, "pieces": 6}
  ],
  "notes": "Brief takeoff assumptions you made"
}

Be thorough — scan every section view, every detail callout, every plan. Do NOT omit small members. If lengths are not explicit, estimate from the drawing scale and shown dimensions."""


# ---------------------------------------------------------------------------
# Helper: fallback weight for sections not in lookup table
# ---------------------------------------------------------------------------
def estimate_section_weight(section: str) -> float:
    m = re.search(r"(\d+)x(\d+)x(\d+)", section or "")
    if not m:
        return 25.0
    thk = int(m.group(3))
    return max(8.0, float(thk) * 1.0)


def get_kg_per_m(section: str) -> tuple[float, bool]:
    """Return (kg/m, is_estimated). Tries exact lookup, then normalised lookup, then fallback."""
    if section in SECTION_WEIGHTS:
        return SECTION_WEIGHTS[section], False
    # Normalise spacing (e.g. "UC152x152x30" → "UC 152x152x30")
    normalised = re.sub(r"([A-Za-z]+)(\d)", r"\1 \2", section).strip()
    if normalised in SECTION_WEIGHTS:
        return SECTION_WEIGHTS[normalised], False
    return estimate_section_weight(section), True


# ---------------------------------------------------------------------------
# Step 1: Call Claude vision API
# ---------------------------------------------------------------------------
async def extract_from_pdf(pdf_bytes: bytes) -> dict:
    """Send PDF to Claude and return parsed JSON extraction."""
    client = anthropic.AsyncAnthropic(api_key=settings.anthropic_api_key)
    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    message = await client.messages.create(
        model=settings.claude_model_drawing,
        max_tokens=4000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    },
                },
                {"type": "text", "text": EXTRACTION_PROMPT},
            ],
        }],
    )

    raw = message.content[0].text if message.content else ""
    return _parse_extraction_response(raw)


def _parse_extraction_response(raw: str) -> dict:
    """Defensively parse Claude's response — strip markdown fences, find first {…}."""
    if not raw or not raw.strip():
        raise ValueError("Empty response from Claude")

    text = raw
    # Strip markdown fences
    m = re.search(r"```json\s*\n(.*?)\n```", text, re.DOTALL)
    if m:
        text = m.group(1)
    else:
        m = re.search(r"```\s*\n(.*?)\n```", text, re.DOTALL)
        if m:
            text = m.group(1)

    # Extract first {...} block
    if not text.strip().startswith("{"):
        m = re.search(r"(\{.*\})", text, re.DOTALL)
        if m:
            text = m.group(1)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try json_repair if available
        try:
            from json_repair import repair_json
            return json.loads(repair_json(text))
        except Exception:
            raise ValueError(f"Could not parse JSON from Claude response: {raw[:500]}")


# ---------------------------------------------------------------------------
# Step 2–4: Compute everything
# ---------------------------------------------------------------------------
def compute_costing(extraction: dict, markup_pct: float = RATIOS["defaultMarkupPct"]) -> dict:
    """
    Given extraction dict (members, plates), compute steel weight, derived
    quantities, all cost line items, overhead, and selling price.
    Returns a flat dict consumed by both the review API and the Excel generator.
    """
    members: List[Dict[str, Any]] = extraction.get("members", [])
    plates:  List[Dict[str, Any]] = extraction.get("plates", [])

    # --- Steel weight ---
    member_rows = []
    total_steel_kg = 0.0
    for m in members:
        section = m.get("section", "")
        length_m = float(m.get("total_length_m") or 0)
        pieces = int(m.get("pieces") or 0)
        role = m.get("role", "")
        kg_per_m, is_estimated = get_kg_per_m(section)
        weight_kg = kg_per_m * length_m
        total_steel_kg += weight_kg
        member_rows.append({
            "section":      section,
            "role":         role,
            "length_m":     round(length_m, 3),
            "kg_per_m":     kg_per_m,
            "pieces":       pieces,
            "weight_kg":    round(weight_kg, 2),
            "is_estimated": is_estimated,
        })

    plate_weight_kg = 0.0
    for p in plates:
        thk = float(p.get("thickness_mm") or 0)
        area = float(p.get("total_area_m2") or 0)
        plate_weight_kg += thk * area * PLATE_DENSITY

    total_steel_kg += plate_weight_kg

    # --- Derived quantities ---
    surface_area_sqm  = round(total_steel_kg * RATIOS["surfaceAreaPerKg"])
    bolts             = math.ceil(total_steel_kg / 1000 * RATIOS["boltsPer1000Kg"])
    paint_litres      = math.ceil(surface_area_sqm * RATIOS["paintLitresPerSqm"])
    mpi_visits        = max(1, math.ceil(total_steel_kg / 1000 * RATIOS["mpiVisitsPer1000Kg"]))
    welding_mh        = math.ceil(total_steel_kg * RATIOS["weldingMHPerKg"])
    fabrication_mh    = math.ceil(total_steel_kg * RATIOS["fabricationMHPerKg"])

    # --- Direct costs ---
    steel_mat_cost    = total_steel_kg * RATIOS["rateSteelPerKg"]
    bolt_cost         = bolts          * RATIOS["rateBoltPerNo"]
    paint_mat_cost    = paint_litres   * RATIOS["ratePaintPerLitre"]
    weld_cost         = welding_mh     * RATIOS["rateWeldingPerMH"]
    fab_cost          = fabrication_mh * RATIOS["rateFabricationPerMH"]
    blast_cost        = surface_area_sqm * RATIOS["rateBlastingPerSqm"]
    paint_app_cost    = surface_area_sqm * RATIOS["ratePaintingPerSqm"]
    mpi_cost          = mpi_visits     * RATIOS["rateMPIPerVisit"]
    qaqc_cost         = RATIOS["rateQAQC"]
    packing_cost      = RATIOS["ratePacking"]

    subtotal_no_consum = (
        steel_mat_cost + bolt_cost + paint_mat_cost +
        weld_cost + fab_cost + blast_cost + paint_app_cost +
        mpi_cost + qaqc_cost + packing_cost
    )

    # --- Overhead (S54) ---
    oh_rate   = 230000 / 30 / 30 / 8
    blast_mh  = (1 / 13.3) * surface_area_sqm * 3
    paint_mh  = (1 / 6.65) * surface_area_sqm * 3
    overhead  = oh_rate * (welding_mh + fabrication_mh + blast_mh + paint_mh)

    # --- Selling price (resolves circular reference) ---
    selling_price = (subtotal_no_consum + overhead) * (1 + markup_pct)
    consumables   = selling_price / 20
    grand_total   = subtotal_no_consum + consumables + overhead
    net_profit    = selling_price - grand_total
    profit_pct    = (net_profit / selling_price) if selling_price else 0.0

    return {
        # Takeoff
        "member_rows":       member_rows,
        "plates":            plates,
        "plate_weight_kg":   round(plate_weight_kg, 2),
        "total_steel_kg":    round(total_steel_kg, 2),
        # Derived quantities
        "surface_area_sqm":  int(surface_area_sqm),
        "bolts":             bolts,
        "paint_litres":      paint_litres,
        "mpi_visits":        mpi_visits,
        "welding_mh":        welding_mh,
        "fabrication_mh":    fabrication_mh,
        # Cost line items
        "steel_mat_cost":    round(steel_mat_cost, 2),
        "bolt_cost":         round(bolt_cost, 2),
        "paint_mat_cost":    round(paint_mat_cost, 2),
        "weld_cost":         round(weld_cost, 2),
        "fab_cost":          round(fab_cost, 2),
        "blast_cost":        round(blast_cost, 2),
        "paint_app_cost":    round(paint_app_cost, 2),
        "mpi_cost":          round(mpi_cost, 2),
        "qaqc_cost":         round(qaqc_cost, 2),
        "packing_cost":      round(packing_cost, 2),
        "subtotal_no_consum": round(subtotal_no_consum, 2),
        "overhead":          round(overhead, 2),
        # Totals
        "consumables":       round(consumables, 2),
        "grand_total":       round(grand_total, 2),
        "selling_price":     round(selling_price, 2),
        "net_profit":        round(net_profit, 2),
        "profit_pct":        round(profit_pct * 100, 2),
        "markup_pct":        round(markup_pct * 100, 2),
    }


# ---------------------------------------------------------------------------
# Step 5: Generate Excel workbook from template
# ---------------------------------------------------------------------------

def _resolve_template_path() -> Path:
    """Resolve the Job Costing Sheet template to an absolute path."""
    configured = Path(settings.drawing_costing_template_path)
    if configured.is_absolute():
        return configured
    # __file__ is backend/app/services/drawing_costing.py → parents[3] = project root
    project_root = Path(__file__).resolve().parents[3]
    return project_root / configured


def generate_excel(costing: dict, project: dict, customer: dict) -> BytesIO:
    """
    Stamp computed values into the Sample Job Costing Sheet.xlsx template.
    All existing formatting, merged cells, borders, and pre-built formulas
    are preserved. Only the variable data cells listed below are written.

    Critical invariants (verified against the template):
      - J33  stays as formula  =D58/20         (already in template)
      - I40  stays as formula  =1/13.3*G40*3   (written here)
      - I41  stays as formula  =1/6.65*G41*3   (written here)
      - S54  stays as formula  =SUM(S15:S43)   (already in template)
      - D56  stays as formula  =S54            (already in template)
      - D58  is HARDCODED to selling_price     (the only number in the footer)
    """
    template_path = _resolve_template_path()
    if not template_path.exists():
        raise FileNotFoundError(
            f"Job Costing Sheet template not found at: {template_path}\n"
            "Set drawing_costing_template_path in your .env or config."
        )

    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active  # template has exactly one sheet

    # Rename sheet to "{jobNo}- Structure Frame"
    job_no = str(customer.get("jobNo") or "XXXX")
    ws.title = f"{job_no}- Structure Frame"[:31]  # Excel max 31 chars

    T   = costing["total_steel_kg"]
    SA  = costing["surface_area_sqm"]
    B   = costing["bolts"]
    PL  = costing["paint_litres"]
    MPI = costing["mpi_visits"]
    WMH = costing["welding_mh"]
    FMH = costing["fabrication_mh"]

    # ----------------------------------------------------------------
    # Header block — only the fields that vary per job
    # ----------------------------------------------------------------
    ws["A3"] = f"REF NO: CNJ/{customer.get('refNo', '')}/01/2025"
    ws["C4"] = customer.get("customerName", "")
    ws["G4"] = date.today()
    ws["O4"] = customer.get("enquiryNo", "")
    ws["G5"] = customer.get("attention", "")
    ws["O5"] = customer.get("jobNo", "")
    ws["G6"] = customer.get("contact", "")

    # ----------------------------------------------------------------
    # Qty / Manhour cells that carry computed values
    # (all K-column formulas and J-column rates already exist in the template)
    # ----------------------------------------------------------------

    # Row 23 — Structural Steel Material
    ws["G23"] = T

    # Row 24 — Bolts
    ws["G24"] = B

    # Row 25 — Paint Material (qty only; B25 formula and J25 already in template)
    ws["G25"] = PL

    # Row 29 — Structural Welding
    ws["G29"] = T
    ws["I29"] = WMH
    ws["M29"] = "2 Welder"

    # Row 30 — Structural Fabrication
    ws["G30"] = T
    ws["I30"] = FMH
    ws["M30"] = "2 Fabricator 2 Helper"

    # Row 32 — Machining (remarks only; qty/MH left blank as in template)
    ws["M32"] = "1  Machinist"

    # Row 33 — Consumables qty (J33 formula =D58/20 already in template)
    ws["G33"] = 1

    # Row 37 — MPI/DPT
    ws["G37"] = MPI

    # Row 40 — Blasting (CRITICAL: I40 must be a formula, not a value)
    ws["G40"] = SA
    ws["I40"] = "=1/13.3*G40*3"
    ws["M40"] = "1 Blaster 1 Helper"

    # Row 41 — Painting (CRITICAL: I41 must be a formula, not a value)
    ws["G41"] = SA
    ws["I41"] = "=1/6.65*G41*3"
    ws["M41"] = "1 Painter 1 Helper"

    # Row 43 — QA/QC qty
    ws["G43"] = 1

    # Row 44 — Packing qty
    ws["G44"] = 1

    # ----------------------------------------------------------------
    # Footer
    # ----------------------------------------------------------------

    # Row 58 — CRITICAL: D58 hardcoded selling price (NUMBER, not formula)
    ws["D58"] = costing["selling_price"]

    # Row 59 — Net Profit formula (cell is empty in the template)
    ws["D59"] = "=D58-D57"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
