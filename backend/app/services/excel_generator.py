"""Excel generator based on the sample workbook template in ReferenceFiles."""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate output Excel in the same format as Sample Job Costing Sheet.xlsx."""

    def __init__(self) -> None:
        # .../backend/app/services/excel_generator.py -> project root at parents[3]
        project_root = Path(__file__).resolve().parents[3]
        self.template_path = project_root / "ReferenceFiles" / "Sample Job Costing Sheet.xlsx"

    def generate(
        self,
        job_data: Dict[str, Any],
        costing_result: Dict[str, Any],
        rates: Dict[str, Any],
    ) -> bytes:
        """Generate workbook bytes using the sample template as base."""
        wb = self._load_template_workbook()
        ws = wb.worksheets[0]

        self._fill_costing_template(ws, job_data, costing_result, rates)
        self._append_rates_sheet(wb, rates)
        self._append_audit_sheet(wb, costing_result.get("audit_trail", []))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _load_template_workbook(self) -> Workbook:
        if not self.template_path.exists():
            raise FileNotFoundError(f"Excel template not found: {self.template_path}")
        return load_workbook(self.template_path)

    def _fill_costing_template(
        self,
        ws,
        job_data: Dict[str, Any],
        costing_result: Dict[str, Any],
        rates: Dict[str, Any],
    ) -> None:
        totals = costing_result or {}
        line_items = totals.get("line_items", []) or []
        extracted_data = job_data.get("extracted_data", {}) or {}

        # ──────────────────────────────────────────────────────────────
        # HEADER SECTION (Rows 1-12)
        # ──────────────────────────────────────────────────────────────
        job_number = str(job_data.get("job_number") or job_data.get("reference_number") or "")
        client_name = str(job_data.get("client_name") or extracted_data.get("client") or "")
        project_name = str(job_data.get("project_name") or extracted_data.get("project_name") or "")
        project_ref = str(job_data.get("project_ref") or extracted_data.get("drawing_number") or "")
        unit_area = str(extracted_data.get("unit_area") or "")

        ws["A3"] = f"REF NO: {job_number}" if job_number else "REF NO:"
        ws["C4"] = client_name
        ws["G4"] = datetime.now().strftime("%Y-%m-%d")
        ws["O4"] = project_ref
        ws["O5"] = job_number
        ws["C7"] = project_name
        ws["G7"] = "X" if ("module" in project_name.lower() or "module" in unit_area.lower()) else ""

        # ──────────────────────────────────────────────────────────────
        # READ COSTING TOTALS — prefer new detailed breakdown fields,
        # fall back to old aggregate fields for backward compatibility
        # ──────────────────────────────────────────────────────────────
        total_weight_kg        = self._safe_float(totals.get("total_weight_kg"))
        total_direct_cost      = self._safe_float(totals.get("total_direct_cost"))
        overhead_cost          = self._safe_float(totals.get("overhead_cost"))
        grand_total            = self._safe_float(totals.get("grand_total"),
                                    (total_direct_cost or 0) + (overhead_cost or 0))
        selling_price          = self._safe_float(totals.get("selling_price"))
        profit_amount          = self._safe_float(totals.get("profit_amount"))

        # Detailed C&J line items (from new breakdown fields)
        bolt_qty        = self._safe_float(totals.get("bolt_qty"), 0.0)
        paint_litres    = self._safe_float(totals.get("paint_litres"), 0.0)
        welding_hrs     = self._safe_float(totals.get("welding_hrs"), 0.0)
        fab_hrs         = self._safe_float(totals.get("fab_hrs"), 0.0)
        blasting_m2     = self._safe_float(totals.get("blasting_m2"), 0.0)
        painting_m2     = self._safe_float(totals.get("painting_m2"), 0.0)
        galv_kg         = self._safe_float(totals.get("galv_kg"), 0.0)
        mpi_visits      = int(self._safe_float(totals.get("mpi_visits"), 1.0))
        mpi_cost        = self._safe_float(totals.get("mpi_cost"), 0.0)
        qaqc_cost       = self._safe_float(totals.get("qaqc_cost"), 3000.0)
        packing_cost    = self._safe_float(totals.get("packing_cost"), 3000.0)

        # Fallback: read from extracted costing_sheet_inputs if totals not populated
        costing_inputs = extracted_data.get("costing_sheet_inputs", {}) or {}
        if total_weight_kg == 0:
            total_weight_kg = self._safe_float(costing_inputs.get("structural_steel_total_kg"))
        if bolt_qty == 0:
            bolt_qty = self._safe_float(costing_inputs.get("bolt_quantity_nos"), 0.0)
        if paint_litres == 0:
            paint_litres = self._safe_float(costing_inputs.get("paint_litres"), 0.0)
        if welding_hrs == 0:
            welding_hrs = self._safe_float(costing_inputs.get("welding_hours"), 0.0)
        if fab_hrs == 0:
            fab_hrs = self._safe_float(costing_inputs.get("fabrication_hours"), 0.0)
        if blasting_m2 == 0:
            blasting_m2 = self._safe_float(costing_inputs.get("blasting_area_m2"), 0.0)
        if painting_m2 == 0:
            painting_m2 = self._safe_float(costing_inputs.get("painting_area_m2"), 0.0)
        if galv_kg == 0:
            galv_kg = self._safe_float(costing_inputs.get("galvanizing_weight_kg"), 0.0)

        # C&J MASTER RATES (AED) — match DEFAULT_RATES in costing_engine.py
        material_rate    = self._safe_float(rates.get("material_rate_per_kg"),              4.00)
        bolt_rate        = self._safe_float(rates.get("bolt_rate_per_nos"),                12.50)
        paint_mat_rate   = self._safe_float(rates.get("paint_material_rate_per_litre"),    21.00)
        welding_rate     = self._safe_float(rates.get("welding_hourly_rate"),              10.50)
        fab_rate         = self._safe_float(rates.get("fabrication_hourly_rate"),           9.50)
        blasting_rate    = self._safe_float(rates.get("blasting_rate_per_m2"),              9.00)
        painting_rate    = self._safe_float(rates.get("painting_rate_per_m2"),             11.00)
        galv_rate        = self._safe_float(rates.get("galvanizing_rate_per_kg"),           2.00)
        mpi_rate         = self._safe_float(rates.get("mpi_rate_per_visit"),              600.00)

        # ──────────────────────────────────────────────────────────────
        # ROW 23: STRUCTURAL STEEL MATERIAL (Kg)
        # ──────────────────────────────────────────────────────────────
        ws["B23"] = "Structural Steel Material"
        ws["G23"] = total_weight_kg if total_weight_kg else ""
        ws["H23"] = "Kg"
        ws["J23"] = material_rate
        # K23 = G23 * J23  (formula preserved in template)

        # ──────────────────────────────────────────────────────────────
        # ROW 24: BOLTS / FASTENERS (Nos)
        # ──────────────────────────────────────────────────────────────
        ws["B24"] = "M20x90 Long Bolts HEX HD"
        ws["G24"] = bolt_qty if bolt_qty > 0 else ""
        ws["H24"] = "Nos"
        ws["J24"] = bolt_rate

        # ──────────────────────────────────────────────────────────────
        # ROW 25: PAINT MATERIAL (litres)
        # ──────────────────────────────────────────────────────────────
        ws["B25"] = "Paint Material"
        ws["G25"] = round(paint_litres, 3) if paint_litres > 0 else ""
        ws["H25"] = "litres"
        ws["J25"] = paint_mat_rate

        # ──────────────────────────────────────────────────────────────
        # ROW 29: WELDING LABOUR (Manhours)
        # ──────────────────────────────────────────────────────────────
        ws["B29"] = "Structural Welding Labour"
        ws["I29"] = round(welding_hrs, 2) if welding_hrs > 0 else ""
        ws["H29"] = "Hrs"
        ws["J29"] = welding_rate

        # ──────────────────────────────────────────────────────────────
        # ROW 30: FABRICATION LABOUR (Manhours)
        # ──────────────────────────────────────────────────────────────
        ws["B30"] = "Structural Fabrication Labour"
        ws["I30"] = round(fab_hrs, 2) if fab_hrs > 0 else ""
        ws["H30"] = "Hrs"
        ws["J30"] = fab_rate

        # ──────────────────────────────────────────────────────────────
        # ROW 39: GALVANIZING (Kg)
        # ──────────────────────────────────────────────────────────────
        ws["B39"] = "Galvanizing Work"
        ws["G39"] = galv_kg if galv_kg > 0 else ""
        ws["H39"] = "Kg"
        ws["J39"] = galv_rate

        # ──────────────────────────────────────────────────────────────
        # ROW 40: BLASTING (SQM)
        # ──────────────────────────────────────────────────────────────
        ws["B40"] = f"Blasting ({round(blasting_m2, 2)} Sq M)" if blasting_m2 > 0 else "Blasting"
        ws["G40"] = round(blasting_m2, 2) if blasting_m2 > 0 else ""
        ws["H40"] = "SQM"
        ws["J40"] = blasting_rate

        # ──────────────────────────────────────────────────────────────
        # ROW 41: PAINTING (SQM)
        # ──────────────────────────────────────────────────────────────
        ws["B41"] = f"Painting ({round(painting_m2, 2)} Sq M)" if painting_m2 > 0 else "Painting"
        ws["G41"] = round(painting_m2, 2) if painting_m2 > 0 else ""
        ws["H41"] = "SQM"
        ws["J41"] = painting_rate

        # ──────────────────────────────────────────────────────────────
        # MPI / DPT INSPECTION — try row 43; write if blank in template
        # ──────────────────────────────────────────────────────────────
        ws["B43"] = f"MPI / DPT Inspection ({mpi_visits} visit{'s' if mpi_visits != 1 else ''})"
        ws["G43"] = mpi_visits if mpi_visits > 0 else ""
        ws["H43"] = "Visits"
        ws["J43"] = mpi_rate

        # ──────────────────────────────────────────────────────────────
        # TOTALS SECTION (Rows 55-62)
        # ──────────────────────────────────────────────────────────────
        # D55: Total Direct Cost
        ws["D55"] = total_direct_cost if total_direct_cost and total_direct_cost > 0 else ""

        # D56: Overheads
        ws["D56"] = overhead_cost if overhead_cost and overhead_cost > 0 else ""

        # D57: GRAND TOTAL (direct + overhead); use costing engine value if present
        if grand_total and grand_total > 0:
            ws["D57"] = grand_total
        elif total_direct_cost and overhead_cost:
            ws["D57"] = round(total_direct_cost + overhead_cost, 2)
        else:
            ws["D57"] = ""

        # D58: Selling Price (margin-on-sell basis)
        ws["D58"] = selling_price if selling_price and selling_price > 0 else ""

        # D59: Net Profit = D58 - D57
        if profit_amount and profit_amount != 0:
            ws["D59"] = profit_amount
        elif selling_price and grand_total:
            ws["D59"] = round(selling_price - grand_total, 2)
        else:
            ws["D59"] = ""

        # F61: Profit Margin % = D59 / D58
        if selling_price and selling_price > 0 and profit_amount:
            ws["F61"] = profit_amount / selling_price
            ws["F61"].number_format = "0.0%"

        # ──────────────────────────────────────────────────────────────
        # QA/QC and Packing rows (fixed costs, append after MPI)
        # ──────────────────────────────────────────────────────────────
        ws["B46"] = "QA/QC Documentation & NDT"
        ws["G46"] = 1
        ws["H46"] = "Lot"
        ws["J46"] = qaqc_cost if qaqc_cost else 3000.0

        ws["B47"] = "Packing & Loading"
        ws["G47"] = 1
        ws["H47"] = "Lot"
        ws["J47"] = packing_cost if packing_cost else 3000.0

        # ──────────────────────────────────────────────────────────────
        # NUMERIC FORMATTING
        # ──────────────────────────────────────────────────────────────
        # Currency format for costs
        for cell_ref in ["D55", "D56", "D57", "D58", "D59"]:
            if ws[cell_ref].value:
                ws[cell_ref].number_format = "#,##0.00"

        # Quantity/Weight format (3 decimals)
        for cell_ref in ["G23", "G24", "G25", "I29", "I30", "G39", "G40", "G41", "G43", "G46", "G47"]:
            if ws[cell_ref].value:
                ws[cell_ref].number_format = "#,##0.000"

        # Rate format (2 decimals)
        for cell_ref in ["J23", "J24", "J25", "J29", "J30", "J40", "J41", "J43", "J46", "J47"]:
            if ws[cell_ref].value:
                ws[cell_ref].number_format = "#,##0.00"

    def _append_rates_sheet(self, wb: Workbook, rates: Dict[str, Any]) -> None:
        ws = wb.create_sheet("RATES & CONFIG")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16

        ws.merge_cells("A1:C1")
        ws["A1"] = "RATE CONFIGURATION SNAPSHOT"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F3864")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Parameter", "Value", "Unit"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(3, col, h)
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = self._thin_border()

        units = {
            "material_rate_per_kg": "AED/kg",
            "fabrication_hourly_rate": "AED/hr",
            "welding_hourly_rate": "AED/hr",
            "surface_treatment_rate_per_m2": "AED/m2",
            "overhead_percentage": "%",
            "profit_margin_percentage": "%",
        }

        row = 4
        for key, val in rates.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, self._safe_float(val, default=0.0))
            ws.cell(row, 3, units.get(key, ""))
            for col in [1, 2, 3]:
                c = ws.cell(row, col)
                c.font = Font(name="Calibri", size=9)
                c.border = self._thin_border()
            ws.cell(row, 2).number_format = "#,##0.000"
            row += 1

    def _append_audit_sheet(self, wb: Workbook, audit_trail: List[Dict[str, Any]]) -> None:
        ws = wb.create_sheet("AUDIT TRAIL")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 80

        ws.merge_cells("A1:C1")
        ws["A1"] = "CALCULATION AUDIT TRAIL"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F3864")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        for idx, h in enumerate(["Step / Item", "Status", "Details"], start=1):
            c = ws.cell(2, idx, h)
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = self._thin_border()

        row = 3
        for entry in audit_trail:
            step = entry.get("item_tag") or entry.get("step") or ""
            status = entry.get("status") or "ok"
            details = {k: v for k, v in entry.items() if k not in {"item_tag", "step", "status"}}
            ws.cell(row, 1, str(step))
            ws.cell(row, 2, str(status))
            ws.cell(row, 3, str(details))
            for col in [1, 2, 3]:
                c = ws.cell(row, col)
                c.font = Font(name="Calibri", size=9)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = self._thin_border()
            row += 1

    @staticmethod
    def _safe_float(value: Any, default: float = 0.0) -> float:
        try:
            if value is None:
                return default
            return float(value)
        except (TypeError, ValueError):
            return default

    def _sum_first_present(self, items: List[Dict[str, Any]], field_names: List[str]) -> float:
        total = 0.0
        for item in items:
            value = None
            for field_name in field_names:
                if field_name in item and item[field_name] is not None:
                    value = item[field_name]
                    break
            if value is not None:
                total += self._safe_float(value, default=0.0)
        return total

    @staticmethod
    def _thin_border() -> Border:
        side = Side(style="thin")
        return Border(left=side, right=side, top=side, bottom=side)


excel_generator = ExcelGenerator()
